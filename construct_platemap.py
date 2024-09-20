import customtkinter as ctk
from pathlib import Path
from openpyxl import Workbook

def get_unique_drugs(pw_combos):
    separate_drugs = [drugs.split('+') for drugs in pw_combos]
    unique_drugs = []

    for sub_list in separate_drugs:
        for drug in sub_list:
            if drug not in unique_drugs:
                unique_drugs.append(drug)

    return unique_drugs
def construct_platemap(unique_drugs, pw_combos, user_inputs):

    desktop_path = Path.home() / "Desktop"
    batch = 28
    r_row = 16
    r_col = 24
    total_drugs = unique_drugs + pw_combos
    batch_data = []

    input_list = [str(word).replace(' ',"").replace('-','') for word in user_inputs]
    exp_type, condition, strain, date = input_list[0], input_list[1], input_list[2], input_list[3]

    # Create batches of data

    for i in range(0, len(total_drugs), batch):
        if i+batch > len(total_drugs):
            batch_data.append(total_drugs[i:])

        else:
            batch_data.append(total_drugs[i:i+batch])

    # Main loop
    for idx, batch in enumerate(batch_data):
        # Create new xlsx file
        wb = Workbook()
        ws = wb.active

    # Ring of blanks
        for row in range(1, r_row + 1):
            for col in range(1, r_col + 1):
                if row == 1 or row == r_row or col == 1 or col == r_col:
                    ws.cell(row=row, column=col, value= "Blank")
    # Last 6 blanks
        for x in range(18,24):
            ws.cell(row=15, column=x, value= "Blank")

    # 2nd row of unt and INHstd
        for x in range(2,10):
            ws.cell(row=2, column=x, value= "unt")
        for x in range(10,24):
            ws.cell(row=2, column=x, value="INHstd")

    # Write cells with drugs
        # Unfilled plate with "Blank"
        batch = batch
        if len(batch) < 28:
            while len(batch) < 28:
                batch.append('Blank')

        ten_batch = [drug for drug in batch for x in range(10)]
        drug_index = 0

        for row in range(3,16):
            for col in range(2,24):
                if drug_index < len(ten_batch):
                    ws.cell(row=row, column=col, value=f"{ten_batch[drug_index]}")
                    drug_index += 1
                else:
                    break

    # One plate in iteration is complete and saved to Desktop
        file_name = f'{date}_DiaMOND_{condition}_{exp_type}_handmadelayout_p{idx+1}.xlsx'
        file_path = desktop_path / file_name
        wb.save(file_path)

if __name__ == "__main__":
    print('hello')
    # unique_drugs = get_unique_drugs(pw_combos)
    #
    # construct_platemap(unique_drugs, pw_combos)



